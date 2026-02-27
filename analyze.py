import openpyxl

wb = openpyxl.load_workbook('FORMULARIO194_MESADEPARTES.xlsx')
print('HOJAS:', wb.sheetnames)

for name in wb.sheetnames:
    ws = wb[name]
    print(f'\n{"="*60}')
    print(f'HOJA: {name}')
    print(f'Filas: {ws.max_row}, Columnas: {ws.max_column}')
    print(f'{"="*60}')
    
    # Headers - row 1
    print('\n--- ENCABEZADOS (Fila 1) ---')
    for c in range(1, min(ws.max_column + 1, 30)):
        val = ws.cell(1, c).value
        if val:
            print(f'  Col {c}: [{val}]')
    
    # Check row 2 for sub-headers
    print('\n--- Fila 2 ---')
    for c in range(1, min(ws.max_column + 1, 30)):
        val = ws.cell(2, c).value
        if val:
            print(f'  Col {c}: [{val}]')
    
    # Check row 3
    print('\n--- Fila 3 ---')
    for c in range(1, min(ws.max_column + 1, 30)):
        val = ws.cell(3, c).value
        if val:
            print(f'  Col {c}: [{val}]')
    
    # Data rows
    print('\n--- DATOS (primeras 5 filas con datos) ---')
    count = 0
    for r in range(4, min(ws.max_row + 1, 50)):
        row_data = []
        for c in range(1, min(ws.max_column + 1, 20)):
            v = ws.cell(r, c).value
            if v is not None:
                row_data.append(f'C{c}={v}')
        if row_data:
            print(f'  Fila {r}: {" | ".join(row_data)}')
            count += 1
            if count >= 5:
                break
    
    # Merged cells
    print(f'\n--- CELDAS COMBINADAS ---')
    for merge in ws.merged_cells.ranges:
        print(f'  {merge}')

print('\n\nANALISIS COMPLETO')
