"""
==========================================================
  SUNAT - Formulario 194 & Expedientes Mesa de Partes
  Modo Claro - Limpio e Intuitivo
==========================================================
"""

import tkinter as tk
from tkinter import messagebox
from datetime import date
import openpyxl
import os

try:
    import xlwings as xw
    HAS_XLWINGS = True
except ImportError:
    HAS_XLWINGS = False

# ============================================================
# COLORES - MODO CLARO
# ============================================================
C = {
    'bg':       '#f5f5f5',
    'white':    '#ffffff',
    'card':     '#ffffff',
    'input_bg': '#ffffff',
    'input_fg': '#1a1a1a',
    'red':      '#c0392b',
    'red_h':    '#e74c3c',
    'red_light':'#fdeaea',
    'blue':     '#2563eb',
    'blue_h':   '#3b82f6',
    'blue_light':'#e8f0fe',
    'green':    '#16a34a',
    'text':     '#1a1a1a',
    'text2':    '#555555',
    'hint':     '#aaaaaa',
    'border':   '#d4d4d4',
    'border_f': '#2563eb',
    'yellow':   '#ca8a04',
    'header_bg':'#c0392b',
}

FONT      = ('Segoe UI', 12)
FONT_B    = ('Segoe UI', 12, 'bold')
FONT_SM   = ('Segoe UI', 10)
FONT_SM_B = ('Segoe UI', 10, 'bold')
FONT_LG   = ('Segoe UI', 18, 'bold')
FONT_MONO = ('Consolas', 13, 'bold')

HEADERS = [
    'N° CAJA', 'N° PAQUETE', 'NUMERO DE REGISTRO', 'TOMO',
    'RANGO INICIAL', 'RANGO FINAL', 'FOLIOS', 'TIPO DOCUMENTAL',
    'N° DE DOCUMENTO', 'RAZON SOCIAL', 'RUC', 'FECHA EXTREMA',
    'OBSERVACIONES', 'X1', 'X2', 'X3'
]

PLACEHOLDERS = {
    'N° DE DOCUMENTO': 'Ej: 1390400006411',
    'RAZON SOCIAL':    'Nombre de la empresa o persona',
    'RUC':             '11 dígitos',
    'FECHA EXTREMA':   'DD/MM/AAAA',
    'OBSERVACIONES':   'Opcional',
}


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("SUNAT - Registro de Formularios")
        self.configure(bg=C['bg'])
        self.resizable(True, True)

        w, h = 520, 700
        x = (self.winfo_screenwidth() - w) // 2
        y = (self.winfo_screenheight() - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")
        self.minsize(450, 620)

        self.excel_dir = os.path.dirname(os.path.abspath(__file__))
        self.excel_path = os.path.join(self.excel_dir, 'FORMULARIO194_MESADEPARTES.xlsx')
        self._ensure_excel()

        self.mode = 'formulario'
        self.entries = {}
        self.exp_parts = []
        self.live_mode = False

        self._build_header()
        self._build_live_indicator()
        self._build_mode_selector()

        # Card principal
        self.card = tk.Frame(self, bg=C['card'], bd=1, relief='solid',
                             highlightbackground=C['border'], highlightthickness=1)
        self.card.pack(fill='both', expand=True, padx=25, pady=(0, 15))

        self.form_frame = tk.Frame(self.card, bg=C['card'])
        self.form_frame.pack(fill='both', expand=True, padx=25, pady=20)

        self._build_form()
        self._build_footer()
        self._update_counter()
        self._check_live()

    # ── EXCEL ──────────────────────────────────────────
    def _ensure_excel(self):
        if os.path.exists(self.excel_path):
            return
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = 'FORMULARIO 194'
        for i, h in enumerate(HEADERS, 1):
            ws1.cell(1, i, h)
        ws2 = wb.create_sheet('EXPEDIENTES')
        for i, h in enumerate(HEADERS, 1):
            ws2.cell(1, i, h)
        wb.save(self.excel_path)

    # ── XLWINGS ────────────────────────────────────────
    def _get_book(self):
        if not HAS_XLWINGS:
            return None
        try:
            fname = os.path.basename(self.excel_path)
            for a in xw.apps:
                for b in a.books:
                    if b.name.lower() == fname.lower():
                        return b
        except Exception:
            pass
        return None

    def _check_live(self):
        was = self.live_mode
        self.live_mode = self._get_book() is not None
        if self.live_mode != was:
            self._update_live_ui()
        self.after(3000, self._check_live)

    def _write_live(self, sheet, row, data):
        book = self._get_book()
        if not book:
            return False
        try:
            ws = book.sheets[sheet]
            for col, val in data.items():
                ws.range((row, col)).value = val
            return True
        except Exception:
            return False

    def _write_offline(self, sheet, row, data):
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb[sheet]
        for col, val in data.items():
            ws.cell(row, col, val)
        wb.save(self.excel_path)
        wb.close()

    def _next_row(self, sheet):
        if self.live_mode:
            book = self._get_book()
            if book:
                try:
                    ws = book.sheets[sheet]
                    if not ws.range('H2').value:
                        return 2
                    return ws.range('H1').end('down').row + 1
                except Exception:
                    pass
        wb = openpyxl.load_workbook(self.excel_path, read_only=True)
        ws = wb[sheet]
        r = ws.max_row + 1 if ws.max_row else 2
        wb.close()
        return r

    def _count(self, sheet):
        if self.live_mode:
            book = self._get_book()
            if book:
                try:
                    ws = book.sheets[sheet]
                    if not ws.range('H2').value:
                        return 0
                    return max(0, ws.range('H1').end('down').row - 1)
                except Exception:
                    pass
        try:
            wb = openpyxl.load_workbook(self.excel_path, read_only=True)
            ws = wb[sheet]
            n = max(0, ws.max_row - 1) if ws.max_row else 0
            wb.close()
            return n
        except Exception:
            return 0

    # ── HEADER ─────────────────────────────────────────
    def _build_header(self):
        hdr = tk.Frame(self, bg=C['header_bg'], height=65)
        hdr.pack(fill='x')
        hdr.pack_propagate(False)

        inner = tk.Frame(hdr, bg=C['header_bg'])
        inner.pack(fill='both', expand=True, padx=25)

        # Logo
        icon = tk.Canvas(inner, width=38, height=38, bg=C['header_bg'], highlightthickness=0)
        icon.pack(side='left', pady=10)
        icon.create_rectangle(2, 2, 36, 36, fill='white', outline='')
        icon.create_text(19, 19, text='S', fill=C['red'], font=('Segoe UI', 18, 'bold'))

        tk.Label(inner, text="  SUNAT", font=FONT_LG, fg='white',
                 bg=C['header_bg']).pack(side='left')

        tk.Label(inner, text="Registro de Formularios", font=FONT_SM,
                 fg='#ffcccc', bg=C['header_bg']).pack(side='right', pady=10)

    # ── INDICADOR LIVE ─────────────────────────────────
    def _build_live_indicator(self):
        self.live_frame = tk.Frame(self, bg=C['bg'])
        self.live_frame.pack(fill='x', padx=25, pady=(10, 0))

        self.live_lbl = tk.Label(self.live_frame, font=FONT_SM, bg=C['bg'])
        self.live_lbl.pack(side='left')
        self._update_live_ui()

    def _update_live_ui(self):
        if self.live_mode:
            self.live_lbl.config(text="🟢 Conectado a Excel — cambios en tiempo real",
                                  fg=C['green'])
        elif HAS_XLWINGS:
            self.live_lbl.config(text="🟡 Abre el Excel para ver cambios en vivo",
                                  fg=C['yellow'])
        else:
            self.live_lbl.config(text="📄 Modo archivo", fg=C['text2'])

    # ── SELECTOR ───────────────────────────────────────
    def _build_mode_selector(self):
        frame = tk.Frame(self, bg=C['bg'])
        frame.pack(fill='x', padx=25, pady=(12, 12))
        frame.columnconfigure(0, weight=1)
        frame.columnconfigure(1, weight=1)

        self.btn_form = tk.Label(frame, text="📋  Formulario 194",
                                 font=FONT_B, fg='white', bg=C['red'],
                                 pady=10, cursor='hand2')
        self.btn_form.grid(row=0, column=0, sticky='nsew', padx=(0, 3))
        self.btn_form.bind('<Button-1>', lambda e: self._set_mode('formulario'))

        self.btn_exp = tk.Label(frame, text="📑  Exp. Mesa de Partes",
                                font=FONT_B, fg=C['text2'], bg='#e5e5e5',
                                pady=10, cursor='hand2')
        self.btn_exp.grid(row=0, column=1, sticky='nsew', padx=(3, 0))
        self.btn_exp.bind('<Button-1>', lambda e: self._set_mode('expediente'))

    def _set_mode(self, mode):
        self.mode = mode
        if mode == 'formulario':
            self.btn_form.config(bg=C['red'], fg='white')
            self.btn_exp.config(bg='#e5e5e5', fg=C['text2'])
        else:
            self.btn_form.config(bg='#e5e5e5', fg=C['text2'])
            self.btn_exp.config(bg=C['blue'], fg='white')
        self._build_form()
        self._update_counter()

    # ── FORMULARIO ─────────────────────────────────────
    def _build_form(self):
        for w in self.form_frame.winfo_children():
            w.destroy()
        self.entries = {}
        self.exp_parts = []

        if self.mode == 'formulario':
            self._add_field('N° DE DOCUMENTO')
            self._add_field('RAZON SOCIAL')
            self._add_field('RUC')
            self._add_field('FECHA EXTREMA')
            self._add_field('OBSERVACIONES')
        else:
            self._add_exp_field()
            self._add_field('RAZON SOCIAL')
            self._add_field('RUC')
            self._add_field('FECHA EXTREMA')
            self._add_field('OBSERVACIONES')

        self._build_buttons()

    def _add_field(self, name):
        hint = PLACEHOLDERS.get(name, '')

        lbl_row = tk.Frame(self.form_frame, bg=C['card'])
        lbl_row.pack(fill='x', pady=(8, 2))
        tk.Label(lbl_row, text=name, font=FONT_SM_B, fg=C['text2'],
                 bg=C['card']).pack(side='left')

        if 'FECHA' in name:
            btn = tk.Label(lbl_row, text="📅 Hoy", font=FONT_SM,
                           fg=C['blue'], bg=C['card'], cursor='hand2')
            btn.pack(side='right')
            btn.bind('<Button-1>', lambda ev, n=name: self._fill_today(n))
            btn.bind('<Enter>', lambda ev, b=btn: b.config(fg=C['blue_h']))
            btn.bind('<Leave>', lambda ev, b=btn: b.config(fg=C['blue']))

        e = tk.Entry(self.form_frame, font=FONT, bg=C['input_bg'], fg=C['hint'],
                     insertbackground=C['text'], relief='solid', bd=1,
                     highlightbackground=C['border'], highlightthickness=1,
                     highlightcolor=C['border_f'])
        e.pack(fill='x', ipady=7)
        e.insert(0, hint)
        e.bind('<FocusIn>', lambda ev, entry=e, h=hint: self._ph_in(entry, h))
        e.bind('<FocusOut>', lambda ev, entry=e, h=hint: self._ph_out(entry, h))
        if name == 'RUC':
            e.bind('<KeyRelease>', lambda ev, entry=e: self._ruc_chk(entry))
        self.entries[name] = e

    def _add_exp_field(self):
        lbl_row = tk.Frame(self.form_frame, bg=C['card'])
        lbl_row.pack(fill='x', pady=(8, 2))
        tk.Label(lbl_row, text="N° DE EXPEDIENTE", font=FONT_SM_B,
                 fg=C['text2'], bg=C['card']).pack(side='left')

        box = tk.Frame(self.form_frame, bg=C['input_bg'], relief='solid', bd=1,
                       highlightbackground=C['border'], highlightthickness=1)
        box.pack(fill='x', ipady=5)

        inner = tk.Frame(box, bg=C['input_bg'])
        inner.pack(padx=10, pady=3)

        tk.Label(inner, text="000-URD", font=FONT_MONO, fg=C['red'],
                 bg=C['input_bg']).pack(side='left')

        self.exp_parts = []
        widths = [3, 4, 6, 1]
        for i, w in enumerate(widths):
            if i > 0:
                tk.Label(inner, text="-", font=FONT_MONO, fg=C['hint'],
                         bg=C['input_bg']).pack(side='left', padx=1)
            e = tk.Entry(inner, font=FONT_MONO, width=w + 1,
                         bg=C['input_bg'], fg=C['text'],
                         insertbackground=C['text'],
                         relief='flat', bd=0, justify='center',
                         highlightthickness=0)
            e.pack(side='left', padx=2)
            self.exp_parts.append(e)
            if i < len(widths) - 1:
                e.bind('<KeyRelease>', lambda ev, idx=i, mw=w: self._exp_tab(ev, idx, mw))

    def _build_buttons(self):
        frame = tk.Frame(self.form_frame, bg=C['card'])
        frame.pack(fill='x', pady=(20, 0))

        color = C['red'] if self.mode == 'formulario' else C['blue']
        hover = C['red_h'] if self.mode == 'formulario' else C['blue_h']

        self.save_btn = tk.Label(frame, text="💾  GUARDAR REGISTRO",
                                 font=FONT_B, fg='white', bg=color,
                                 pady=12, cursor='hand2')
        self.save_btn.pack(fill='x')
        self.save_btn.bind('<Button-1>', lambda e: self._save())
        self.save_btn.bind('<Enter>', lambda e: self.save_btn.config(bg=hover))
        self.save_btn.bind('<Leave>', lambda e: self.save_btn.config(bg=color))

        clr = tk.Label(frame, text="🔄 Limpiar", font=FONT_SM,
                       fg=C['text2'], bg=C['card'], pady=6, cursor='hand2')
        clr.pack(fill='x', pady=(6, 0))
        clr.bind('<Button-1>', lambda e: self._clear())
        clr.bind('<Enter>', lambda e: clr.config(fg=C['text']))
        clr.bind('<Leave>', lambda e: clr.config(fg=C['text2']))

    # ── FOOTER ─────────────────────────────────────────
    def _build_footer(self):
        ft = tk.Frame(self, bg='#e8e8e8', height=35)
        ft.pack(fill='x', side='bottom')
        ft.pack_propagate(False)

        self.count_lbl = tk.Label(ft, text="", font=FONT_SM, fg=C['text2'], bg='#e8e8e8')
        self.count_lbl.pack(side='left', padx=15, pady=6)

        self.status_lbl = tk.Label(ft, text="", font=FONT_SM, fg=C['green'], bg='#e8e8e8')
        self.status_lbl.pack(side='right', padx=15, pady=6)

    def _update_counter(self):
        sheet = 'FORMULARIO 194' if self.mode == 'formulario' else 'EXPEDIENTES'
        n = self._count(sheet)
        name = "Formulario 194" if self.mode == 'formulario' else "Expedientes"
        self.count_lbl.config(text=f"📊 {n} registros en {name}")

    # ── HELPERS ────────────────────────────────────────
    def _ph_in(self, entry, hint):
        if entry.get() == hint:
            entry.delete(0, 'end')
            entry.config(fg=C['input_fg'])

    def _ph_out(self, entry, hint):
        if not entry.get().strip():
            entry.delete(0, 'end')
            entry.insert(0, hint)
            entry.config(fg=C['hint'])

    def _val(self, name):
        e = self.entries.get(name)
        if not e: return ''
        v = e.get().strip()
        return '' if v == PLACEHOLDERS.get(name, '') else v

    def _fill_today(self, name):
        e = self.entries.get(name)
        if e:
            e.delete(0, 'end')
            e.insert(0, date.today().strftime('%d/%m/%Y'))
            e.config(fg=C['input_fg'])

    def _ruc_chk(self, entry):
        v = entry.get().strip()
        if v == PLACEHOLDERS.get('RUC', ''):
            entry.config(highlightbackground=C['border'], highlightthickness=1)
        elif len(v) == 11 and v.isdigit():
            entry.config(highlightbackground=C['green'], highlightthickness=2)
        elif v:
            entry.config(highlightbackground=C['red'], highlightthickness=2)

    def _exp_tab(self, ev, idx, mw):
        if len(self.exp_parts[idx].get()) >= mw:
            self.exp_parts[idx + 1].focus_set()

    def _clear(self):
        for name, e in self.entries.items():
            e.delete(0, 'end')
            e.insert(0, PLACEHOLDERS.get(name, ''))
            e.config(fg=C['hint'], highlightbackground=C['border'], highlightthickness=1)
        for e in self.exp_parts:
            e.delete(0, 'end')

    def _shake(self, name):
        e = self.entries.get(name)
        if e:
            e.config(highlightbackground=C['red'], highlightthickness=2)
            self.after(1500, lambda: e.config(highlightbackground=C['border'], highlightthickness=1))

    def _flash(self):
        if hasattr(self, 'save_btn'):
            color = C['red'] if self.mode == 'formulario' else C['blue']
            self.save_btn.config(bg=C['green'], text="✅  ¡GUARDADO!")
            self.after(1200, lambda: self.save_btn.config(bg=color, text="💾  GUARDAR REGISTRO"))

    def _status(self, txt, err=False):
        self.status_lbl.config(text=txt, fg=C['red'] if err else C['green'])
        self.after(4000, lambda: self.status_lbl.config(text=""))

    # ── SAVE ───────────────────────────────────────────
    def _save(self):
        try:
            if self.mode == 'formulario':
                self._save_f194()
            else:
                self._save_exp()
        except Exception as ex:
            self._status(f"❌ {ex}", err=True)
            messagebox.showerror("Error", str(ex))

    def _save_f194(self):
        doc   = self._val('N° DE DOCUMENTO')
        razon = self._val('RAZON SOCIAL')
        ruc   = self._val('RUC')
        fecha = self._val('FECHA EXTREMA')
        obs   = self._val('OBSERVACIONES')

        if not doc:
            self._shake('N° DE DOCUMENTO')
            return messagebox.showwarning("Atención", "Ingrese el N° de Documento")
        if not ruc or len(ruc) != 11 or not ruc.isdigit():
            self._shake('RUC')
            return messagebox.showwarning("Atención", "El RUC debe tener 11 dígitos")

        sheet = 'FORMULARIO 194'
        row = self._next_row(sheet)
        data = {8:'FORMULARIO NRO. 194', 9:doc, 10:razon, 11:ruc, 12:fecha, 13:obs}

        live = self.live_mode and self._write_live(sheet, row, data)
        if not live:
            self._write_offline(sheet, row, data)

        self._clear()
        self._update_counter()
        self._status("✅ Formulario guardado" + (" ⚡" if live else ""))
        self._flash()

    def _save_exp(self):
        parts = [e.get().strip() for e in self.exp_parts]
        if not all(parts):
            if self.exp_parts: self.exp_parts[0].focus_set()
            return messagebox.showwarning("Atención", "Complete el N° de Expediente")

        doc   = f"000-URD{parts[0]}-{parts[1]}-{parts[2]}-{parts[3]}"
        razon = self._val('RAZON SOCIAL')
        ruc   = self._val('RUC')
        fecha = self._val('FECHA EXTREMA')
        obs   = self._val('OBSERVACIONES')

        if not ruc or len(ruc) != 11 or not ruc.isdigit():
            self._shake('RUC')
            return messagebox.showwarning("Atención", "El RUC debe tener 11 dígitos")

        sheet = 'EXPEDIENTES'
        row = self._next_row(sheet)
        data = {8:'EXPEDIENTE DE MEZA DE PARTES', 9:doc, 10:razon, 11:ruc, 12:fecha, 13:obs}

        live = self.live_mode and self._write_live(sheet, row, data)
        if not live:
            self._write_offline(sheet, row, data)

        self._clear()
        self._update_counter()
        self._status("✅ Expediente guardado" + (" ⚡" if live else ""))
        self._flash()


if __name__ == '__main__':
    app = App()
    app.mainloop()
