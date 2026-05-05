"""
SUNAT - Sistema de Registro de Formularios
Backend FastAPI con soporte xlwings (tiempo real) + openpyxl (offline)
"""

from fastapi import FastAPI, Request
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel, field_validator
import openpyxl
import os
import webbrowser
import threading
import httpx

# xlwings opcional
try:
    import xlwings as xw
    HAS_XLWINGS = True
except ImportError:
    HAS_XLWINGS = False

# ─── CONFIG ────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, "FORMULARIO194_MESADEPARTES.xlsx")

HEADERS = [
    'N° CAJA', 'N° PAQUETE', 'NUMERO DE REGISTRO', 'TOMO',
    'RANGO INICIAL', 'RANGO FINAL', 'FOLIOS', 'TIPO DOCUMENTAL',
    'N° DE DOCUMENTO', 'RAZON SOCIAL', 'RUC', 'FECHA EXTREMA',
    'OBSERVACIONES', 'X1', 'X2', 'X3'
]

app = FastAPI(title="SUNAT Formularios")
app.mount("/static", StaticFiles(directory=os.path.join(BASE_DIR, "static")), name="static")
templates = Jinja2Templates(directory=os.path.join(BASE_DIR, "templates"))


# ─── MODELS ────────────────────────────────────────────
class RegistroFormulario(BaseModel):
    documento: str
    razon_social: str = ""
    ruc: str
    fecha: str = ""
    observaciones: str = ""

    @field_validator("ruc")
    @classmethod
    def ruc_must_be_11(cls, v):
        v = v.strip()
        if len(v) != 11 or not v.isdigit():
            raise ValueError("RUC debe tener exactamente 11 dígitos")
        return v


class RegistroExpediente(BaseModel):
    parte1: str
    parte2: str
    parte3: str
    parte4: str
    razon_social: str = ""
    ruc: str
    fecha: str = ""
    observaciones: str = ""

    @field_validator("ruc")
    @classmethod
    def ruc_must_be_11(cls, v):
        v = v.strip()
        if len(v) != 11 or not v.isdigit():
            raise ValueError("RUC debe tener exactamente 11 dígitos")
        return v


# ─── EXCEL HELPERS ─────────────────────────────────────
def ensure_excel():
    if os.path.exists(EXCEL_PATH):
        return
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "FORMULARIO 194"
    for i, h in enumerate(HEADERS, 1):
        ws1.cell(1, i, h)
    ws2 = wb.create_sheet("EXPEDIENTES")
    for i, h in enumerate(HEADERS, 1):
        ws2.cell(1, i, h)
    wb.save(EXCEL_PATH)


def get_xlwings_book():
    if not HAS_XLWINGS:
        return None
    try:
        fname = os.path.basename(EXCEL_PATH)
        for a in xw.apps:
            for b in a.books:
                if b.name.lower() == fname.lower():
                    return b
    except Exception:
        pass
    return None


def write_record(sheet_name: str, data: dict) -> dict:
    """Escribe un registro. Intenta xlwings primero, luego openpyxl."""
    live = False
    book = get_xlwings_book()

    if book:
        try:
            ws = book.sheets[sheet_name]
            if not ws.range("H2").value:
                row = 2
            else:
                row = ws.range("H1").end("down").row + 1
            for col, val in data.items():
                ws.range((row, col)).value = val
            live = True
        except Exception:
            live = False

    if not live:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb[sheet_name]
        row = ws.max_row + 1
        for col, val in data.items():
            ws.cell(row, col, val)
        wb.save(EXCEL_PATH)
        wb.close()

    return {"success": True, "live": live, "row": row}


def count_records(sheet_name: str) -> int:
    book = get_xlwings_book()
    if book:
        try:
            ws = book.sheets[sheet_name]
            if not ws.range("H2").value:
                return 0
            return max(0, ws.range("H1").end("down").row - 1)
        except Exception:
            pass
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
        ws = wb[sheet_name]
        n = max(0, ws.max_row - 1) if ws.max_row else 0
        wb.close()
        return n
    except Exception:
        return 0


# ─── ROUTES ────────────────────────────────────────────
@app.on_event("startup")
def startup():
    ensure_excel()


@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


@app.get("/api/status")
async def api_status():
    book = get_xlwings_book()
    return {
        "excel_connected": book is not None,
        "has_xlwings": HAS_XLWINGS,
        "formulario_count": count_records("FORMULARIO 194"),
        "expediente_count": count_records("EXPEDIENTES"),
    }


@app.get("/api/ruc/{ruc}")
async def get_ruc_info(ruc: str):
    url = f"https://app1.dirislimacentro.gob.pe/std/mod_ext/api_sunat.php?action=qry_sunat_ruc&ruc={ruc}"
    async with httpx.AsyncClient(timeout=10.0) as client:
        try:
            response = await client.get(url)
            if response.status_code == 200:
                return response.json()
            return JSONResponse(status_code=response.status_code, content={"error": "API error"})
        except Exception as e:
            return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/api/formulario")
async def save_formulario(reg: RegistroFormulario):
    fecha_val = reg.fecha.strip()
    if fecha_val:
        fecha_val = f"'{fecha_val}"

    data = {
        8: "FORMULARIO NRO. 194",
        9: reg.documento.strip(),
        10: reg.razon_social.strip(),
        11: reg.ruc.strip(),
        12: fecha_val,
        13: reg.observaciones.strip(),
    }
    result = write_record("FORMULARIO 194", data)
    result["count"] = count_records("FORMULARIO 194")
    return result


@app.post("/api/expediente")
async def save_expediente(reg: RegistroExpediente):
    fecha_val = reg.fecha.strip()
    if fecha_val:
        fecha_val = f"'{fecha_val}"

    doc = f"000-URD{reg.parte1.strip()}-{reg.parte2.strip()}-{reg.parte3.strip()}-{reg.parte4.strip()}"
    data = {
        8: "EXPEDIENTE DE MEZA DE PARTES",
        9: doc,
        10: reg.razon_social.strip(),
        11: reg.ruc.strip(),
        12: fecha_val,
        13: reg.observaciones.strip(),
    }
    result = write_record("EXPEDIENTES", data)
    result["count"] = count_records("EXPEDIENTES")
    return result


# ─── RUN ───────────────────────────────────────────────
if __name__ == "__main__":
    import uvicorn
    threading.Timer(1.5, lambda: webbrowser.open("http://127.0.0.1:8000")).start()
    uvicorn.run(app, host="127.0.0.1", port=8000)
